import pandas as pd
import numpy as np
from datetime import datetime
import os

class ExcelHandler:
    def __init__(self):
        self.required_columns = [
            'patient_id', 'cart_date', 'date', 'anc',
            'last_fu_date', 'subsequent_therapy_date', 'progression_date'
        ]
    
    def validate_file(self, filepath):
        """Validate uploaded Excel file structure"""
        try:
            # Try to read the file
            df = pd.read_excel(filepath)
            
            if len(df) == 0:
                return {'valid': False, 'message': 'File is empty'}
            
            # Check for required columns
            missing_cols = [col for col in self.required_columns if col not in df.columns]
            if missing_cols:
                return {
                    'valid': False,
                    'message': f'Missing required columns: {", ".join(missing_cols)}'
                }
            
            # Check for patient IDs
            if df['patient_id'].isna().all():
                return {'valid': False, 'message': 'No valid patient IDs found'}
            
            # Check for date columns
            date_columns = ['cart_date', 'date']
            for col in date_columns:
                try:
                    pd.to_datetime(df[col], errors='coerce')
                except:
                    return {'valid': False, 'message': f'Invalid date format in column: {col}'}
            
            # Check for ANC values
            try:
                pd.to_numeric(df['anc'], errors='coerce')
            except:
                return {'valid': False, 'message': 'Invalid ANC values - must be numeric'}
            
            return {'valid': True, 'message': 'File validation successful'}
            
        except Exception as e:
            return {'valid': False, 'message': f'File reading error: {str(e)}'}
    
    def load_data(self, filepath):
        """Load and clean data from Excel file"""
        df = pd.read_excel(filepath)
        
        # Convert date columns
        date_columns = ['cart_date', 'date', 'last_fu_date', 'subsequent_therapy_date', 'progression_date']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Convert ANC to numeric
        df['anc'] = pd.to_numeric(df['anc'], errors='coerce')
        
        # Remove rows with missing essential data
        df = df.dropna(subset=['patient_id', 'cart_date'])
        
        return df
    
    def export_results(self, results_df, output_buffer):
        """Export results to Excel with formatting"""
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # Main results sheet
            results_df.to_excel(writer, sheet_name='ICAHT_Grades', index=False)
            
            # Summary sheet
            summary_data = self._create_summary_sheet(results_df)
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format worksheets
            self._format_worksheets(writer, results_df)
    
    def _create_summary_sheet(self, results_df):
        """Create summary statistics sheet"""
        summary_data = []
        
        # Early ICAHT summary
        early_counts = results_df['early_icaht_grade'].value_counts().sort_index()
        for grade, count in early_counts.items():
            summary_data.append({
                'Category': 'Early ICAHT',
                'Grade': grade,
                'Count': count,
                'Percentage': f"{(count/len(results_df)*100):.1f}%"
            })
        
        # Late ICAHT summary
        late_counts = results_df['late_icaht_grade'].value_counts().sort_index()
        for grade, count in late_counts.items():
            summary_data.append({
                'Category': 'Late ICAHT',
                'Grade': grade,
                'Count': count,
                'Percentage': f"{(count/len(results_df)*100):.1f}%"
            })
        
        # Special cases
        grade_4_special = results_df['grade_4_special'].sum() if 'grade_4_special' in results_df.columns else 0
        summary_data.append({
            'Category': 'Special Cases',
            'Grade': 'Grade 4 (Never recovered)',
            'Count': grade_4_special,
            'Percentage': f"{(grade_4_special/len(results_df)*100):.1f}%"
        })
        
        return pd.DataFrame(summary_data)
    
    def _format_worksheets(self, writer, results_df):
        """Apply formatting to Excel worksheets"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format main results sheet
        ws_results = writer.sheets['ICAHT_Grades']
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws_results[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # Grade color coding
        grade_colors = {
            'Grade 0': PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid"),
            'Grade 1': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'Grade 2': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
            'Grade 3': PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
            'Grade 4': PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        }
        
        # Apply grade colors
        for row in ws_results.iter_rows(min_row=2):
            early_grade = row[1].value if len(row) > 1 else None
            late_grade = row[2].value if len(row) > 2 else None
            
            if early_grade in grade_colors:
                row[1].fill = grade_colors[early_grade]
            if late_grade in grade_colors:
                row[2].fill = grade_colors[late_grade]
